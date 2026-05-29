"""Generate sample Excel upload templates with openpyxl."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1E2D45")
HEADER_FONT = Font(name="Calibri", bold=True, color="E8C97A", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="2A3A55"),
    right=Side(style="thin", color="2A3A55"),
    top=Side(style="thin", color="2A3A55"),
    bottom=Side(style="thin", color="2A3A55"),
)
DOB_NUMBER_FORMAT = "DD-MM-YYYY"


def _autosize_columns(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_sheet(ws, headers: list[str], sample_row: list, *, date_columns: set[int] | None = None) -> None:
    date_columns = date_columns or set()

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for col_idx, value in enumerate(sample_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if col_idx in date_columns:
            cell.number_format = DOB_NUMBER_FORMAT

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def build_employee_sample_workbook() -> BytesIO:
    """Build sample employee master Excel template in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Master"

    headers = ["Employee ID", "Name", "Email", "Designation", "Date of Birth"]
    sample_row = [
        "EMP001",
        "Arjun Menon",
        "arjun@example.com",
        "Engineer",
        date(1995, 6, 15),
    ]
    _write_sheet(ws, headers, sample_row, date_columns={5})
    _autosize_columns(ws, {1: 14, 2: 22, 3: 28, 4: 18, 5: 16})

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_salary_sample_workbook() -> BytesIO:
    """Build sample monthly salary Excel template in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Sheet"

    headers = [
        "Employee ID",
        "Base Salary",
        "HRA",
        "Allowances",
        "Deductions",
        "Month",
        "Year",
    ]
    sample_row = ["EMP001", 25000, 5000, 2000, 1500, "May", 2026]
    _write_sheet(ws, headers, sample_row)

    for col_idx in range(2, 6):
        ws.cell(row=2, column=col_idx).number_format = "#,##0.00"

    _autosize_columns(ws, {1: 14, 2: 14, 3: 10, 4: 14, 5: 14, 6: 12, 7: 10})

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
